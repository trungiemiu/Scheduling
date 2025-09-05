# cplex_solver/cplex_solver.py
# Python 3.7 • IBM CPLEX 12.9 • docplex.mp
# FJSSP with Job-Splitting (Cmax), includes setup (optional) and minimum sub-lot size s
# Robust reader: filters out non-positive IDs and rows whose job is not in Lotsize.

from __future__ import annotations
import argparse
from collections import defaultdict
from typing import Dict, Tuple, List, Set
import sys, os

try:
    import openpyxl
except ImportError:
    raise SystemExit("Please `pip install openpyxl` for Excel reading.")
try:
    from docplex.mp.model import Model
except ImportError:
    raise SystemExit("Please ensure `docplex` is installed and CPLEX is available.")

# ------------------------------ Data loading ------------------------------

def read_excel_data(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def read_table(wsname: str):
        if wsname not in wb.sheetnames:
            return None, []
        ws = wb[wsname]
        header = None
        rows = []
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r == 0:
                header = [str(c).strip().lower() if c is not None else "" for c in row]
                continue
            if all(c is None for c in row):
                continue
            rows.append(row)
        return header, rows

    # --- Lotsize ---
    ls_name = next((n for n in wb.sheetnames if n.strip().lower() == "lotsize"), None)
    if not ls_name:
        raise ValueError("Missing 'Lotsize' sheet.")
    header, rows = read_table(ls_name)
    colmap = {c: i for i, c in enumerate(header)}
    jcol = colmap.get("job")
    gcol = colmap.get("g") or colmap.get("lotsize") or colmap.get("lot size")
    if jcol is None or gcol is None:
        # fallback: assume first col is job, next numeric col is g
        jcol = 0 if jcol is None else jcol
        if gcol is None:
            for idx in range(len(header)):
                if idx == jcol:
                    continue
                if any(isinstance(r[idx], (int, float)) for r in rows if r[idx] is not None):
                    gcol = idx
                    break
    if gcol is None:
        raise ValueError("Cannot detect Lotsize value column.")
    g: Dict[int, int] = {}
    for r in rows:
        jv, gv = r[jcol], r[gcol]
        if jv is None or gv is None:
            continue
        try:
            j = int(jv)
            val = int(gv)
        except Exception:
            continue
        if j <= 0 or val < 0:
            continue  # filter non-positive/invalid
        g[j] = val

    valid_jobs = {j for j in g.keys() if j > 0}
    if not valid_jobs:
        raise ValueError("No valid jobs in Lotsize.")

    # --- Processing time ---
    pt_sheet = None
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "") in ("processingtime", "processing_time", "processing"):
            pt_sheet = name
            break
    if pt_sheet is None:
        raise ValueError("Missing 'ProcessingTime' sheet.")
    header, rows = read_table(pt_sheet)
    if not rows:
        raise ValueError("ProcessingTime is empty.")
    hmap = {c: i for i, c in enumerate(header)}
    cj = hmap.get("job")
    co = hmap.get("operation") or hmap.get("op")
    cm = hmap.get("machine") or hmap.get("m")
    cp = (hmap.get("processing time") or hmap.get("processingtime") or
          hmap.get("time") or hmap.get("p"))
    if cj is None or co is None or cm is None or cp is None:
        raise ValueError(f"ProcessingTime must have (job, operation, machine, p). Headers: {header}")

    p_all: Dict[Tuple[int, int, int], float] = {}
    for r in rows:
        jv, ov, mv, pv = r[cj], r[co], r[cm], r[cp]
        if jv is None or ov is None or mv is None or pv is None:
            continue
        try:
            i, j, k = int(jv), int(ov), int(mv)
            val = float(pv)
        except Exception:
            continue
        if i <= 0 or j <= 0 or k <= 0:
            continue  # ignore non-positive IDs
        p_all[(i, j, k)] = val

    # keep only rows whose job is present in Lotsize
    p: Dict[Tuple[int, int, int], float] = {t: v for t, v in p_all.items() if t[0] in valid_jobs}
    if not p:
        raise ValueError("No (job,op,machine) in ProcessingTime match Jobs in Lotsize.")

    # --- Setup time (optional) ---
    h_all: Dict[Tuple[int, int, int], float] = defaultdict(float)
    st_sheet = next((n for n in wb.sheetnames
                    if n.strip().lower().replace(" ", "") in ("setuptime", "setup")), None)
    if st_sheet:
        header, rows = read_table(st_sheet)
        if rows:
            hmap = {c: i for i, c in enumerate(header)}
            cj2 = hmap.get("job")
            co2 = hmap.get("operation") or hmap.get("op")
            cm2 = hmap.get("machine") or hmap.get("m")
            ch2 = hmap.get("h") or hmap.get("setup") or hmap.get("setup time")
            if cj2 is not None and co2 is not None and cm2 is not None and ch2 is not None:
                for r in rows:
                    jv, ov, mv, hv = r[cj2], r[co2], r[cm2], r[ch2]
                    if jv is None or ov is None or mv is None or hv is None:
                        continue
                    try:
                        i, j, k = int(jv), int(ov), int(mv)
                        val = float(hv)
                    except Exception:
                        continue
                    if i <= 0 or j <= 0 or k <= 0:
                        continue
                    if i in valid_jobs:
                        h_all[(i, j, k)] = val
    h = h_all  # already filtered by valid jobs where applicable

    # --- JOSet / JOMSet (optional) ---
    def load_named(name):
        if name in wb.sheetnames:
            return read_table(name)
        return (None, [])

    header_JO, rows_JO = load_named("JOSet")
    header_JOM, rows_JOM = load_named("JOMSet")

    if rows_JO:
        cmap = {c: idx for idx, c in enumerate(header_JO)}
        cj_ = cmap.get("job")
        co_ = cmap.get("op") or cmap.get("operation")
        if cj_ is None or co_ is None:
            raise ValueError("JOSet must have [job, op].")
        JOSet_raw = []
        for r in rows_JO:
            if r[cj_] is None or r[co_] is None:
                continue
            try:
                i, j = int(r[cj_]), int(r[co_])
            except Exception:
                continue
            if i <= 0 or j <= 0:
                continue
            JOSet_raw.append((i, j))
        # keep only jobs that exist in Lotsize
        JOSet = [(i, j) for (i, j) in JOSet_raw if i in valid_jobs]
        if not JOSet:
            # fallback to infer from p
            JOSet = sorted({(i, j) for (i, j, _) in p.keys()})
    else:
        JOSet = sorted({(i, j) for (i, j, _) in p.keys()})

    if rows_JOM:
        cmap = {c: idx for idx, c in enumerate(header_JOM)}
        cj_ = cmap.get("job")
        co_ = cmap.get("op") or cmap.get("operation")
        cm_ = cmap.get("machine")
        if cj_ is None or co_ is None or cm_ is None:
            raise ValueError("JOMSet must have [job, op, machine].")
        JOM_raw = []
        for r in rows_JOM:
            if r[cj_] is None or r[co_] is None or r[cm_] is None:
                continue
            try:
                i, j, k = int(r[cj_]), int(r[co_]), int(r[cm_])
            except Exception:
                continue
            if i <= 0 or j <= 0 or k <= 0:
                continue
            JOM_raw.append((i, j, k))
        # keep only jobs in Lotsize and also present in p
        JOMSet = [(i, j, k) for (i, j, k) in JOM_raw if i in valid_jobs and (i, j, k) in p]
        if not JOMSet:
            JOMSet = sorted(list(p.keys()))
    else:
        JOMSet = sorted(list(p.keys()))

    # Build indexes consistent with Lotsize
    I: List[int] = sorted({i for (i, _) in JOSet if i in valid_jobs})
    from collections import defaultdict as dd
    J_by_i: Dict[int, List[int]] = dd(list)
    for i, j in JOSet:
        if i in valid_jobs and j > 0:
            J_by_i[i].append(j)
    for i in list(J_by_i.keys()):
        J_by_i[i] = sorted(set(J_by_i[i]))

    K: List[int] = sorted({k for (_, _, k) in JOMSet if k > 0})
    elig: Set[Tuple[int, int, int]] = {(i, j, k) for (i, j, k) in JOMSet if i in valid_jobs and j > 0 and k > 0}

    # Validate (only over filtered sets)
    for (i, j, k) in elig:
        if (i, j, k) not in p:
            raise ValueError(f"Missing p_ijk for eligible triple ({i},{j},{k})")
    # I is already subset of g's jobs; this check is now tautological but keep it safe:
    for i in I:
        if i not in g:
            raise ValueError(f"Missing Lotsize for job {i}")

    return I, J_by_i, K, elig, g, p, h

# ------------------------------ Model (unchanged) ------------------------------

def build_and_solve(xlsx_path: str, s_min: int = 0, timelimit: int = 600, mipgap: float = 0.0, log_output: bool = True):
    I, J_by_i, K, elig, g, p, h = read_excel_data(xlsx_path)

    delta_by_k = {k: max(1, sum(1 for (i, j, kk) in elig if kk == k)) for k in K}
    sum_hp = sum(h.get((i, j, k), 0.0) + p[(i, j, k)] * g[i] for (i, j, k) in elig)
    M = max(1.0, sum_hp)

    mdl = Model(name="FJSSP_JS_Cmax")

    x = {(i,j,k,l): mdl.binary_var(name=f"x_{i}_{j}_{k}_{l}")
         for (i,j,k) in elig for l in range(1, delta_by_k[k]+1)}
    alpha = {(i,j,k): mdl.integer_var(lb=0, name=f"a_{i}_{j}_{k}") for (i,j,k) in elig}
    C = {(i,j,k): mdl.continuous_var(lb=0, name=f"C_{i}_{j}_{k}") for (i,j,k) in elig}
    Cmax = mdl.continuous_var(lb=0, name="Cmax")

    mdl.minimize(Cmax)

    for (i,j,k) in elig:
        mdl.add_constraint(Cmax >= C[(i,j,k)])

    for i in I:
        for j in J_by_i[i]:
            mdl.add_constraint(mdl.sum(alpha[(ii,jj,kk)] for (ii,jj,kk) in elig if ii==i and jj==j) == g[i])

    for (i,j,k) in elig:
        sum_x = mdl.sum(x[(i,j,k,l)] for l in range(1, delta_by_k[k]+1))
        mdl.add_constraint(alpha[(i,j,k)] <= g[i] * sum_x)
        if s_min > 0:
            mdl.add_constraint(alpha[(i,j,k)] >= s_min * sum_x)

    for k in K:
        for l in range(1, delta_by_k[k]+1):
            mdl.add_constraint(mdl.sum(x[(i,j,k,l)] for (i,j,kk) in elig if kk==k) <= 1)

    for (i,j,k) in elig:
        mdl.add_constraint(mdl.sum(x[(i,j,k,l)] for l in range(1, delta_by_k[k]+1)) <= 1)

    for k in K:
        for l in range(2, delta_by_k[k]+1):
            mdl.add_constraint(
                mdl.sum(x[(i,j,k,l)] for (i,j,kk) in elig if kk==k) -
                mdl.sum(x[(i,j,k,l-1)] for (i,j,kk) in elig if kk==k) <= 0
            )

    for (i,j,k) in elig:
        hij = h.get((i,j,k), 0.0); pij = p[(i,j,k)]
        if (i,j,k,1) in x:
            mdl.add_constraint(C[(i,j,k)] + M*(1 - x[(i,j,k,1)]) >= hij + pij*alpha[(i,j,k)])

    for k in K:
        for l in range(2, delta_by_k[k]+1):
            for (i,j,kk) in elig:
                if kk != k: continue
                hij = h.get((i,j,k), 0.0); pij = p[(i,j,k)]
                for (w,q,kk2) in elig:
                    if kk2 != k: continue
                    mdl.add_constraint(
                        C[(i,j,k)] + M*(2 - x[(i,j,k,l)] - x[(w,q,k,l-1)]) >=
                        C[(w,q,k)] + hij + pij*alpha[(i,j,k)]
                    )

    for i in I:
        ops = J_by_i[i]
        for idx in range(1, len(ops)):
            j = ops[idx]; jprev = ops[idx-1]
            for (ii,jj,k) in elig:
                if not (ii==i and jj==j): continue
                hij = h.get((i,j,k), 0.0); pij = p[(i,j,k)]
                for (iii,jjj,a) in elig:
                    if not (iii==i and jjj==jprev): continue
                    for l in range(1, delta_by_k[k]+1):
                        for b in range(1, delta_by_k[a]+1):
                            mdl.add_constraint(
                                C[(i,j,k)] + M*(2 - x[(i,j,k,l)] - x[(i,jprev,a,b)]) >=
                                C[(i,jprev,a)] + hij + pij*alpha[(i,j,k)]
                            )

    for (i,j,k) in elig:
        hij = h.get((i,j,k), 0.0)
        if hij > 0:
            sum_x = mdl.sum(x[(i,j,k,l)] for l in range(1, delta_by_k[k]+1))
            mdl.add_constraint(C[(i,j,k)] >= hij * sum_x)

    if timelimit is not None and timelimit > 0:
        mdl.context.cplex_parameters.timelimit = int(timelimit)
    if mipgap is not None and mipgap > 0:
        mdl.parameters.mip.tolerances.mipgap = float(mipgap)

    sol = mdl.solve(log_output=log_output)
    if not sol:
        print("No solution found."); return None

    Cmax_val = sol.get_value(Cmax)
    print("\n=== CPLEX Solution summary ===")
    print("Cmax =", round(Cmax_val, 4))

    for k in K:
        seq = []
        for l in range(1, delta_by_k[k]+1):
            chosen = [(i,j) for (i,j,kk) in elig if kk==k and sol.get_value(x[(i,j,k,l)]) > 0.5]
            if not chosen: break
            i,j = chosen[0]
            a_val = sol.get_value(alpha[(i,j,k)])
            C_val = sol.get_value(C[(i,j,k)])
            seq.append((l,i,j,a_val,C_val))
        if seq:
            print("\nMachine k = {}:".format(k))
            print("  pos | job | op | sublot | completion")
            for l,i,j,a_val,C_val in seq:
                print("  {:>3} | {:>3} | {:>2} | {:>6.2f} | {:>9.2f}".format(l,i,j,a_val,C_val))
    return sol

# ------------------------------ CLI ------------------------------

def main():
    parser = argparse.ArgumentParser(description="FJSSP-JS (Cmax) with CPLEX 12.9")
    parser.add_argument("--xlsx", type=str, default="data/2data.xlsx")
    parser.add_argument("--smin", type=int, default=0)
    parser.add_argument("--timelimit", type=int, default=600)
    parser.add_argument("--mipgap", type=float, default=0.0)
    parser.add_argument("--no-log", action="store_true")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print("Excel file not found:", args.xlsx); sys.exit(1)

    build_and_solve(args.xlsx, s_min=args.smin, timelimit=args.timelimit,
                    mipgap=args.mipgap, log_output=(not args.no_log))

if __name__ == "__main__":
    main()
